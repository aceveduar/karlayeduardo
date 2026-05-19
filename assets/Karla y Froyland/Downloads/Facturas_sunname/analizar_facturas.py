"""
Script para analizar facturas CFDI (XML del SAT) y generar reporte en Excel.
Detecta duplicados y agrupa facturación por mes.

USO:
    python analizar_facturas.py "C:/ruta/a/tu/carpeta/de/facturas"

El reporte se guardará en la misma carpeta con el nombre "Reporte_Facturacion.xlsx"
"""

import os
import sys
import xml.etree.ElementTree as ET
from datetime import datetime
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Namespaces de CFDI
NAMESPACES = {
    'cfdi4': 'http://www.sat.gob.mx/cfd/4',
    'cfdi3': 'http://www.sat.gob.mx/cfd/3',
    'tfd': 'http://www.sat.gob.mx/TimbreFiscalDigital'
}

def parse_cfdi(filepath):
    """Extrae información de un archivo XML CFDI."""
    try:
        tree = ET.parse(filepath)
        root = tree.getroot()
        
        # Detectar versión (CFDI 4.0 o 3.3)
        ns = root.tag.split('}')[0] + '}' if '}' in root.tag else ''
        
        # Datos del comprobante
        version = root.get('Version') or root.get('version', '')
        fecha_str = root.get('Fecha') or root.get('fecha', '')
        total = float(root.get('Total') or root.get('total', 0))
        subtotal = float(root.get('SubTotal') or root.get('subTotal', 0))
        tipo_comprobante = root.get('TipoDeComprobante') or root.get('tipoDeComprobante', '')
        moneda = root.get('Moneda') or root.get('moneda', 'MXN')
        folio = root.get('Folio') or root.get('folio', '')
        serie = root.get('Serie') or root.get('serie', '')
        
        # Datos del emisor
        emisor = root.find(f'{ns}Emisor')
        if emisor is None:
            emisor = root.find('.//cfdi4:Emisor', NAMESPACES) or root.find('.//cfdi3:Emisor', NAMESPACES)
        
        rfc_emisor = emisor.get('Rfc') or emisor.get('rfc', '') if emisor is not None else ''
        nombre_emisor = emisor.get('Nombre') or emisor.get('nombre', '') if emisor is not None else ''
        
        # Datos del receptor
        receptor = root.find(f'{ns}Receptor')
        if receptor is None:
            receptor = root.find('.//cfdi4:Receptor', NAMESPACES) or root.find('.//cfdi3:Receptor', NAMESPACES)
        
        rfc_receptor = receptor.get('Rfc') or receptor.get('rfc', '') if receptor is not None else ''
        nombre_receptor = receptor.get('Nombre') or receptor.get('nombre', '') if receptor is not None else ''
        
        # UUID del timbre fiscal
        uuid = ''
        for tfd in root.iter():
            if 'TimbreFiscalDigital' in tfd.tag:
                uuid = tfd.get('UUID') or tfd.get('uuid', '')
                break
        
        # Parsear fecha
        fecha = None
        if fecha_str:
            try:
                fecha = datetime.fromisoformat(fecha_str.replace('Z', '+00:00'))
            except:
                try:
                    fecha = datetime.strptime(fecha_str[:19], '%Y-%m-%dT%H:%M:%S')
                except:
                    pass
        
        return {
            'uuid': uuid.upper(),
            'fecha': fecha,
            'fecha_str': fecha_str,
            'total': total,
            'subtotal': subtotal,
            'tipo': tipo_comprobante,
            'moneda': moneda,
            'folio': folio,
            'serie': serie,
            'rfc_emisor': rfc_emisor,
            'nombre_emisor': nombre_emisor,
            'rfc_receptor': rfc_receptor,
            'nombre_receptor': nombre_receptor,
            'version': version,
            'archivo': os.path.basename(filepath)
        }
        
    except Exception as e:
        return {'error': str(e), 'archivo': os.path.basename(filepath)}

def analizar_facturas(carpeta):
    """Analiza todas las facturas XML en una carpeta."""
    facturas = []
    errores = []
    uuids_vistos = {}
    duplicados = []
    
    # Buscar archivos XML
    archivos_xml = [f for f in os.listdir(carpeta) if f.lower().endswith('.xml')]
    
    print(f"Encontrados {len(archivos_xml)} archivos XML")
    
    for archivo in archivos_xml:
        filepath = os.path.join(carpeta, archivo)
        datos = parse_cfdi(filepath)
        
        if 'error' in datos:
            errores.append(datos)
            continue
        
        # Verificar duplicados por UUID
        if datos['uuid']:
            if datos['uuid'] in uuids_vistos:
                duplicados.append({
                    'uuid': datos['uuid'],
                    'archivo_original': uuids_vistos[datos['uuid']],
                    'archivo_duplicado': archivo,
                    'total': datos['total']
                })
                continue  # No agregar duplicado a la lista principal
            else:
                uuids_vistos[datos['uuid']] = archivo
        
        # Solo incluir facturas de ingreso (tipo 'I') - lo que facturaste
        if datos['tipo'] == 'I':
            facturas.append(datos)
    
    return facturas, duplicados, errores

def crear_reporte_excel(facturas, duplicados, errores, carpeta_salida):
    """Genera el reporte en Excel."""
    wb = Workbook()
    
    # Estilos
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='1F4E79')
    money_format = '$#,##0.00'
    date_format = 'DD/MM/YYYY'
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ===== HOJA 1: RESUMEN POR MES =====
    ws_resumen = wb.active
    ws_resumen.title = 'Resumen por Mes'
    
    # Agrupar por mes
    por_mes = defaultdict(lambda: {'total': 0, 'cantidad': 0})
    for f in facturas:
        if f['fecha']:
            mes_key = f['fecha'].strftime('%Y-%m')
            mes_nombre = f['fecha'].strftime('%B %Y')
            por_mes[mes_key]['total'] += f['total']
            por_mes[mes_key]['cantidad'] += 1
            por_mes[mes_key]['nombre'] = mes_nombre
    
    # Headers resumen
    headers_resumen = ['Mes', 'Cantidad de Facturas', 'Total Facturado']
    for col, header in enumerate(headers_resumen, 1):
        cell = ws_resumen.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    # Datos por mes (ordenados)
    row = 2
    meses_ordenados = sorted(por_mes.keys())
    for mes_key in meses_ordenados:
        datos = por_mes[mes_key]
        # Convertir mes a español
        fecha_mes = datetime.strptime(mes_key, '%Y-%m')
        meses_es = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
                    'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
        mes_nombre = f"{meses_es[fecha_mes.month - 1]} {fecha_mes.year}"
        
        ws_resumen.cell(row=row, column=1, value=mes_nombre).border = border
        ws_resumen.cell(row=row, column=2, value=datos['cantidad']).border = border
        ws_resumen.cell(row=row, column=2).alignment = Alignment(horizontal='center')
        cell_total = ws_resumen.cell(row=row, column=3, value=datos['total'])
        cell_total.number_format = money_format
        cell_total.border = border
        row += 1
    
    # Fila de totales
    row_total = row
    ws_resumen.cell(row=row_total, column=1, value='TOTAL').font = Font(bold=True)
    ws_resumen.cell(row=row_total, column=1).border = border
    ws_resumen.cell(row=row_total, column=2, value=f'=SUM(B2:B{row-1})').border = border
    ws_resumen.cell(row=row_total, column=2).alignment = Alignment(horizontal='center')
    ws_resumen.cell(row=row_total, column=2).font = Font(bold=True)
    cell_gran_total = ws_resumen.cell(row=row_total, column=3, value=f'=SUM(C2:C{row-1})')
    cell_gran_total.number_format = money_format
    cell_gran_total.font = Font(bold=True)
    cell_gran_total.border = border
    
    # Ajustar anchos
    ws_resumen.column_dimensions['A'].width = 20
    ws_resumen.column_dimensions['B'].width = 22
    ws_resumen.column_dimensions['C'].width = 20
    
    # ===== HOJA 2: DETALLE DE FACTURAS =====
    ws_detalle = wb.create_sheet('Detalle Facturas')
    
    headers_detalle = ['Fecha', 'Serie', 'Folio', 'UUID', 'Cliente (RFC)', 
                       'Cliente (Nombre)', 'Subtotal', 'Total', 'Moneda', 'Archivo']
    for col, header in enumerate(headers_detalle, 1):
        cell = ws_detalle.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    # Ordenar facturas por fecha
    facturas_ordenadas = sorted(facturas, key=lambda x: x['fecha'] or datetime.min)
    
    for row, f in enumerate(facturas_ordenadas, 2):
        ws_detalle.cell(row=row, column=1, value=f['fecha']).number_format = date_format
        ws_detalle.cell(row=row, column=2, value=f['serie'])
        ws_detalle.cell(row=row, column=3, value=f['folio'])
        ws_detalle.cell(row=row, column=4, value=f['uuid'])
        ws_detalle.cell(row=row, column=5, value=f['rfc_receptor'])
        ws_detalle.cell(row=row, column=6, value=f['nombre_receptor'])
        ws_detalle.cell(row=row, column=7, value=f['subtotal']).number_format = money_format
        ws_detalle.cell(row=row, column=8, value=f['total']).number_format = money_format
        ws_detalle.cell(row=row, column=9, value=f['moneda'])
        ws_detalle.cell(row=row, column=10, value=f['archivo'])
        
        for col in range(1, 11):
            ws_detalle.cell(row=row, column=col).border = border
    
    # Ajustar anchos detalle
    anchos_detalle = [12, 8, 10, 38, 15, 35, 15, 15, 10, 40]
    for i, ancho in enumerate(anchos_detalle, 1):
        ws_detalle.column_dimensions[get_column_letter(i)].width = ancho
    
    # ===== HOJA 3: DUPLICADOS =====
    if duplicados:
        ws_dup = wb.create_sheet('Duplicados Detectados')
        
        headers_dup = ['UUID', 'Archivo Original', 'Archivo Duplicado', 'Monto']
        for col, header in enumerate(headers_dup, 1):
            cell = ws_dup.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = PatternFill('solid', fgColor='C00000')  # Rojo para alertar
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        for row, dup in enumerate(duplicados, 2):
            ws_dup.cell(row=row, column=1, value=dup['uuid']).border = border
            ws_dup.cell(row=row, column=2, value=dup['archivo_original']).border = border
            ws_dup.cell(row=row, column=3, value=dup['archivo_duplicado']).border = border
            cell_monto = ws_dup.cell(row=row, column=4, value=dup['total'])
            cell_monto.number_format = money_format
            cell_monto.border = border
        
        ws_dup.column_dimensions['A'].width = 38
        ws_dup.column_dimensions['B'].width = 40
        ws_dup.column_dimensions['C'].width = 40
        ws_dup.column_dimensions['D'].width = 15
    
    # ===== HOJA 4: ERRORES (si hay) =====
    if errores:
        ws_err = wb.create_sheet('Errores')
        ws_err.cell(row=1, column=1, value='Archivo').font = header_font
        ws_err.cell(row=1, column=1).fill = PatternFill('solid', fgColor='FFA500')
        ws_err.cell(row=1, column=2, value='Error').font = header_font
        ws_err.cell(row=1, column=2).fill = PatternFill('solid', fgColor='FFA500')
        
        for row, err in enumerate(errores, 2):
            ws_err.cell(row=row, column=1, value=err['archivo'])
            ws_err.cell(row=row, column=2, value=err['error'])
        
        ws_err.column_dimensions['A'].width = 40
        ws_err.column_dimensions['B'].width = 60
    
    # Guardar
    output_path = os.path.join(carpeta_salida, 'Reporte_Facturacion.xlsx')
    wb.save(output_path)
    return output_path

def main():
    if len(sys.argv) < 2:
        carpeta = os.path.dirname(os.path.abspath(__file__))
        print(f"Usando carpeta del script: {carpeta}")
    else:
        carpeta = sys.argv[1]
    
    if not os.path.isdir(carpeta):
        print(f"Error: La carpeta '{carpeta}' no existe.")
        sys.exit(1)
    
    print("=" * 60)
    print("ANALIZANDO FACTURAS...")
    print("=" * 60)
    
    facturas, duplicados, errores = analizar_facturas(carpeta)
    
    print(f"\n[OK] Facturas procesadas: {len(facturas)}")
    print(f"[!]  Duplicados encontrados: {len(duplicados)}")
    if errores:
        print(f"[X]  Archivos con errores: {len(errores)}")

    if facturas:
        output = crear_reporte_excel(facturas, duplicados, errores, carpeta)
        print(f"\n[OK] Reporte generado: {output}")
        
        # Mostrar resumen rápido
        total_facturado = sum(f['total'] for f in facturas)
        print(f"\n{'=' * 60}")
        print(f"TOTAL FACTURADO: ${total_facturado:,.2f} MXN")
        print(f"{'=' * 60}")
    else:
        print("\nNo se encontraron facturas de ingreso (emitidas) para procesar.")

if __name__ == '__main__':
    main()
